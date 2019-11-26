from decimal import Decimal
import os
import openpyxl
from openpyxl.styles import Font, colors, Alignment, Border, Side, NamedStyle
from settlements.report.convert_str_contain_comma_v2 import convert_str_contain_comma_v2
BASE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),'DAY_AAA')
COPY_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),'DAY_BBB')

def create_day_excel(stub_data):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    # 设置title块的样式
    title_style = NamedStyle(name='title_style')
    title_style.font = Font(color=colors.BLACK)

    bd = Side(border_style='thin', color='FF000000')
    title_style.border = Border(left=bd,
                                top=bd,
                                right=bd,
                                bottom=bd
                                )

    # horizontal 水平     vertical 垂直
    title_style.alignment = Alignment(horizontal='center',
                                      vertical='center'
                                      )

    # 设置content块的样式
    content_style1 = NamedStyle(name='content_style1')
    content_style1.font = Font(color=colors.BLACK)

    bd = Side(border_style='thin', color='FF000000')
    content_style1.border = Border(left=bd,
                                   top=bd,
                                   right=bd,
                                   bottom=bd
                                   )

    # horizontal 水平     vertical 垂直
    content_style1.alignment = Alignment(horizontal='right',
                                         vertical='center'
                                         )

    # 设置合计的样式
    content_style2 = NamedStyle(name='content_style2')
    content_style2.font = Font(color=colors.BLACK)

    bd = Side(border_style='thin', color='FF000000')
    content_style2.border = Border(left=bd,
                                   top=bd,
                                   right=bd,
                                   bottom=bd
                                   )

    # horizontal 水平     vertical 垂直
    content_style2.alignment = Alignment(horizontal='left',
                                         vertical='center'
                                         )

    # 设置合计求和的样式
    content_style3 = NamedStyle(name='content_style3')
    content_style3.font = Font(color=colors.BLACK)

    bd = Side(border_style='thin', color='FF000000')
    content_style3.border = Border(left=bd,
                                   top=bd,
                                   right=bd,
                                   bottom=bd
                                   )

    # horizontal 水平     vertical 垂直
    content_style3.alignment = Alignment(horizontal='right',
                                         vertical='center'
                                         )


    # 创建命名样式后，将其注册到工作簿中
    wb.add_named_style(title_style)
    wb.add_named_style(content_style1)
    wb.add_named_style(content_style2)
    wb.add_named_style(content_style3)


    title_index_list = ['A1', 'B1', 'C1', 'D1', 'D2', 'E2', 'F2', 'G1', 'G2', 'H2', 'I2']
    title_value_list = ['日付', '加盟店数', '店舗数', '取引件数', 'WeChat', 'Alipay', 'WeChat+Alipay',
                        '取扱高', 'WeChat', 'Alipay', 'WeChat+Alipay', ]

    # title块赋值及样式设计
    for t, t_v in zip(title_index_list, title_value_list):
        ws1[t].style = title_style
        ws1[t].value = t_v

    # 合并单元格
    ws1.merge_cells('A1:A2')
    ws1.merge_cells('B1:B2')
    ws1.merge_cells('C1:C2')
    ws1.merge_cells('D1:F1')
    ws1.merge_cells('G1:I1')


    ws1.style = title_style


    data_list = []
    # 加盟店数
    master_mch_total_count = 0
    # 店铺数
    sub_mch_total_count = 0
    # 取引件数
    transaction_total_count_wechat = 0
    transaction_total_count_alipay = 0
    transaction_total_count_wechat_alipay = 0
    # 去扱高
    transaction_fee_wechat_count = 0
    transaction_fee_alipay_count = 0
    transaction_fee_wechat_alipay_count = 0

    for day_data in stub_data:
        list_ = []
        date = day_data['date'].replace('-','.')
        list_.append(date)
        master_mch_count = day_data['master_mch_count']
        master_mch_total_count = master_mch_count
        list_.append(master_mch_count)

        sub_mch_count = day_data['sub_mch_count']
        sub_mch_total_count = sub_mch_count
        list_.append(sub_mch_count)
        # 获取是否开通微信 未开通则显示-
        use_wechat = day_data['use_wechat']
        if use_wechat == 'Y':
            transaction_count_wechat = day_data['transaction_count']['wechat']
            transaction_total_count_wechat += transaction_count_wechat
        else:
            transaction_count_wechat = '-'
            transaction_total_count_wechat += 0
        list_.append(transaction_count_wechat)
        #获取是否开通支付宝 未开通则显示-
        use_alipay = day_data['use_alipay']
        if use_alipay == 'Y':
            transaction_count_alipay = day_data['transaction_count']['alipay']
            transaction_total_count_alipay += transaction_count_alipay
        else:
            transaction_count_alipay = '-'
            transaction_total_count_alipay += 0
        list_.append(transaction_count_alipay)

        transaction_count_wechat_alipay = day_data['transaction_count']['wechat+alipay']
        transaction_total_count_wechat_alipay += transaction_count_wechat_alipay
        list_.append(transaction_count_wechat_alipay)
        if use_wechat == 'Y':
            transaction_fee_wechat = day_data['transaction_fee']['wechat']
            transaction_fee_wechat_count += transaction_fee_wechat
            new_transaction_fee_wechat = convert_str_contain_comma_v2(transaction_fee_wechat)
        else:
            new_transaction_fee_wechat = '-'
            transaction_fee_wechat_count += 0
        list_.append(new_transaction_fee_wechat)
        if use_alipay == 'Y':
            transaction_fee_alipay = day_data['transaction_fee']['alipay']
            transaction_fee_alipay_count += transaction_fee_alipay
            new_transaction_fee_alipay = convert_str_contain_comma_v2(transaction_fee_alipay)
        else:
            new_transaction_fee_alipay = '-'
            transaction_fee_alipay_count += 0
        list_.append(new_transaction_fee_alipay)

        transaction_fee_wechat_alipay = day_data['transaction_fee']['wechat+alipay']
        transaction_fee_wechat_alipay_count += transaction_fee_wechat_alipay
        new_transaction_fee_wechat_alipay = convert_str_contain_comma_v2(transaction_fee_wechat_alipay)
        list_.append(new_transaction_fee_wechat_alipay)
        data_list.append(list_)
# print(data_list)

    # 写入每日所有数据
    for i, item in enumerate(data_list):
        i = i + 3
        for j, val in enumerate(item):
            ws1.cell(row=i, column=j+1, value=val).style = content_style1

    # 设置行高
    ws1.row_dimensions[1].height = 18
    # 设置列宽
    column_name_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    column_width_list = [13, 18, 18, 10, 10, 18, 20, 20, 20]
    for c, l in zip(column_name_list, column_width_list):
        ws1.column_dimensions[c].width = l


    # 最后一行加入新的一行
    ws1.insert_rows(ws1.max_row + 1)
    row = ws1.max_row + 1
    ws1['A' + str(row)].value = '合計'
    ws1['A' + str(row)].style = content_style2

    columns_list = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    for columns in range(3, row+1):
        for c in columns_list:
            ws1[c + str(columns)].style = content_style1

    ws1['B' + str(row)].value = master_mch_total_count
    ws1['C' + str(row)].value = sub_mch_total_count
    ws1['D' + str(row)].value = transaction_total_count_wechat
    ws1['E' + str(row)].value = transaction_total_count_alipay
    ws1['F' + str(row)].value = transaction_total_count_wechat_alipay
    ws1['G' + str(row)].value = transaction_fee_wechat_count
    ws1['H' + str(row)].value = transaction_fee_alipay_count
    ws1['I' + str(row)].value = transaction_fee_wechat_alipay_count

    # 当前日报表的最后一天为报表标题
    print(data_list)
    file_name = data_list[-1][0]
    file_name = file_name.replace('.','')
    finally_file_name = '日報_Rakuten_' + file_name + '.xlsx'
    wb.save(os.path.join(BASE_DIR, finally_file_name))
    os.system('cp %s %s' % (os.path.join(BASE_DIR, finally_file_name), os.path.join(COPY_DIR, finally_file_name)))
