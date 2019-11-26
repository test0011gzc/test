import os
import datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, colors, Alignment, Border, Side, NamedStyle
from settlements.report.convert_str_contain_comma_v2 import convert_str_contain_comma_v2
from openpyxl import load_workbook
BASE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),'AAA')
COPY_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),'BBB')
'''
Font            设置文字的大小，颜色和下划线等
PatternFill     填充图案和渐变色
Border          单元格的边框
Alignment       单元格的对齐方式等
Protection      写保护
'''
# stub_data = {
#         'total_count': 12,
#         # 'begin_date': '2019-09-01',
#         # 'end_date': '2019-09-15',
#         'begin_date': '2019-09-16',
#         'end_date': '2019-09-30',
#         'data': [{
#             'agency_id': 10114,
#             'agency_name': 'skaura',
#             'count': 2,
#             'data_list': [{
#                 'master_mch_id': 20551,
#                 'master_mch_name': 'グヒュスデュフゲ',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('1.50'),
#                     'alipay': Decimal('1.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 2,
#                     'alipay': 0,
#                     'wechat+alipay': 2
#                 },
#                 'transaction_fee': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 },
#                 'agency_profit': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 }
#             }, {
#                 'master_mch_id': 21048,
#                 'master_mch_name': '',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('1.00'),
#                     'alipay': Decimal('0.80')
#                 },
#                 'transaction_count': {
#                     'wechat': 4,
#                     'alipay': 6,
#                     'wechat+alipay': 10
#                 },
#                 'transaction_fee': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 },
#                 'agency_profit': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 }
#             }]
#         }, {
#             'agency_id': 10078,
#             'agency_name': '乐天银行',
#             'count': 2,
#             'data_list': [{
#                 'master_mch_id': 20806,
#                 'master_mch_name': '',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('2.00'),
#                     'alipay': Decimal('3.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 4,
#                     'alipay': 8,
#                     'wechat+alipay': 12
#                 },
#                 'transaction_fee': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 },
#                 'agency_profit': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 }
#             }, {
#                 'master_mch_id': 27649,
#                 'master_mch_name': 'kshertest002',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('1.00'),
#                     'alipay': Decimal('1.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 6,
#                     'alipay': 10,
#                     'wechat+alipay': 16
#                 },
#                 'transaction_fee': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 },
#                 'agency_profit': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 }
#             }]
#         }, {
#             'agency_id': 10109,
#             'agency_name': 'amy再次回归',
#             'count': 5,
#             'data_list': [{
#                 'master_mch_id': 20943,
#                 'master_mch_name': '',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('2.00'),
#                     'alipay': Decimal('3.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 1,
#                     'alipay': 0,
#                     'wechat+alipay': 1
#                 },
#                 'transaction_fee': {
#                     'wechat': 100,
#                     'alipay': 0,
#                     'wechat+alipay': 100
#                 },
#                 'agency_profit': {
#                     'wechat': 1,
#                     'alipay': 0,
#                     'wechat+alipay': 1
#                 }
#             }, {
#                 'master_mch_id': 20945,
#                 'master_mch_name': '',
#                 'sub_mch_count': 2,
#                 'mdr': {
#                     'wechat': Decimal('2.00'),
#                     'alipay': Decimal('3.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 2,
#                     'alipay': 4,
#                     'wechat+alipay': 6
#                 },
#                 'transaction_fee': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 },
#                 'agency_profit': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 }
#             }, {
#                 'master_mch_id': 20948,
#                 'master_mch_name': '',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('2.00'),
#                     'alipay': Decimal('3.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 2,
#                     'alipay': 6,
#                     'wechat+alipay': 8
#                 },
#                 'transaction_fee': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 },
#                 'agency_profit': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 }
#             }, {
#                 'master_mch_id': 20952,
#                 'master_mch_name': '',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('3.00'),
#                     'alipay': Decimal('1.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 3,
#                     'alipay': 5,
#                     'wechat+alipay': 8
#                 },
#                 'transaction_fee': {
#                     'wechat': 100,
#                     'alipay': 100,
#                     'wechat+alipay': 200
#                 },
#                 'agency_profit': {
#                     'wechat': 1,
#                     'alipay': 0,
#                     'wechat+alipay': 1
#                 }
#             }, {
#                 'master_mch_id': 20950,
#                 'master_mch_name': '',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('2.00'),
#                     'alipay': Decimal('3.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 0,
#                     'alipay': 2,
#                     'wechat+alipay': 2
#                 },
#                 'transaction_fee': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 },
#                 'agency_profit': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 }
#             }]
#         }, {
#             'agency_id': 10093,
#             'agency_name': 'vip sakura',
#             'count': 3,
#             'data_list': [{
#                 'master_mch_id': 20960,
#                 'master_mch_name': '',
#                 'sub_mch_count': 2,
#                 'mdr': {
#                     'wechat': Decimal('1.90'),
#                     'alipay': Decimal('1.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 18,
#                     'alipay': 22,
#                     'wechat+alipay': 40
#                 },
#                 'transaction_fee': {
#                     'wechat': 200,
#                     'alipay': 200,
#                     'wechat+alipay': 400
#                 },
#                 'agency_profit': {
#                     'wechat': 1,
#                     'alipay': 1,
#                     'wechat+alipay': 2
#                 }
#             }, {
#                 'master_mch_id': 21053,
#                 'master_mch_name': '',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('0.90'),
#                     'alipay': Decimal('1.00')
#                 },
#                 'transaction_count': {
#                     'wechat': 14,
#                     'alipay': 24,
#                     'wechat+alipay': 38
#                 },
#                 'transaction_fee': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 },
#                 'agency_profit': {
#                     'wechat': 0,
#                     'alipay': 0,
#                     'wechat+alipay': 0
#                 }
#             }, {
#                 'master_mch_id': 21084,
#                 'master_mch_name': '',
#                 'sub_mch_count': 1,
#                 'mdr': {
#                     'wechat': Decimal('1.00'),
#                     'alipay': Decimal('0.80')
#                 },
#                 'transaction_count': {
#                     'wechat': 11,
#                     'alipay': 15,
#                     'wechat+alipay': 26
#                 },
#                 'transaction_fee': {
#                     'wechat': 100787895645.88,
#                     'alipay': 100,
#                     'wechat+alipay': 200
#                 },
#                 'agency_profit': {
#                     'wechat': 1,
#                     'alipay': 0,
#                     'wechat+alipay': 1
#                 }
#             }]
#         }]
#     }



# 获取所有数据 加入到list_中
# stub_data = {
#  'total_count': 1,
#  'begin_date': '2019-09-01',
#  'end_date': '2019-09-15',
#  'data': [{
#   'agency_id': 10114,
#   'agency_name': 'skaura',
#   'count': 1,
#   'data_list': [{
#    'master_mch_id': 20551,
#    'master_mch_name': 'グヒュスデュフゲ',
#    'sub_mch_count': 3,
#    'mdr': {
#     'wechat': Decimal('1.50'),
#     'alipay': Decimal('1.00')
#    },
#    'transaction_count': {
#     'wechat': 2,
#     'alipay': 0,
#     'wechat+alipay': 2
#    },
#    'transaction_fee': {
#     'wechat': 0,
#     'alipay': 0,
#     'wechat+alipay': 0
#    },
#    'agency_profit': {
#     'wechat': 0,
#     'alipay': 0,
#     'wechat+alipay': 0
#    }
#   }]
#  }]
# }

def create_week_excel(stub_data):

    list_ = []
    # 店铺数
    sub_mch_total_count = 0
    # 取引件数
    transaction_total_count_wechat = 0
    transaction_total_count_alipay = 0
    transaction_total_count_wechat_alipay = 0
    # 取扱高
    transaction_fee_wechat_count = 0
    transaction_fee_alipay_count = 0
    transaction_fee_wechat_alipay_count = 0
    # 银行收益
    agency_profit_wechat_count = 0
    agency_profit_alipay_count = 0
    agency_profit_wechat_alipay_count = 0

    sub_count = stub_data['data']
    for sub in sub_count:
        info = sub['data_list']
        for i in info:
            data_list = []
            # print(i)
            # 代理店名
            agency_name = sub['agency_name']
            # 加盟店名
            master_mch_name = i['master_mch_name']
            # print(master_mch_name)
            data_list.append(master_mch_name)

            # sub_mch_count = i.setdefault('sub_mch_count')
            sub_mch_count = i['sub_mch_count']
            sub_mch_total_count += sub_mch_count
            data_list.append(sub_mch_count)

            # mdr wechat
            mdr_wechat = i['mdr']['wechat']
            mdr_wechat_format = mdr_wechat if mdr_wechat not in [-1,'-1'] else '-'
            data_list.append(str(mdr_wechat_format))

            # mdr alipay
            mdr_alipay = i['mdr']['alipay']
            mdr_alipay_format = mdr_alipay if mdr_alipay not in [-1,'-1'] else '-'
            data_list.append(str(mdr_alipay_format))
            
            # 取引件数 wechat   # 有空值  需要判断
            transaction_count_wechat = i['transaction_count']['wechat'] if mdr_wechat not in [-1,'-1'] else '-'
            if transaction_count_wechat != '-':
                transaction_total_count_wechat += transaction_count_wechat
            else:
                transaction_total_count_wechat += 0
            data_list.append(transaction_count_wechat)
            # 取引件数 alipay
            transaction_count_alipay = i['transaction_count']['alipay'] if mdr_alipay not in [-1,'-1'] else '-'
            if transaction_count_alipay != '-':
                transaction_total_count_alipay += transaction_count_alipay
            else:
                transaction_total_count_alipay += 0
            data_list.append(transaction_count_alipay)
            # 取引件数 wechat+alipay
            transaction_count_wechat_alipay = i['transaction_count']['wechat+alipay']
            transaction_total_count_wechat_alipay += transaction_count_wechat_alipay
            data_list.append(transaction_count_wechat_alipay)

            # 取扱高 wechat
            transaction_fee_wechat = i['transaction_fee']['wechat'] if mdr_wechat not in [-1,'-1'] else '-'
            if transaction_fee_wechat != '-':
                transaction_fee_wechat_count += transaction_fee_wechat
            else:
                transaction_fee_wechat_count += 0
            new_transaction_fee_wechat = convert_str_contain_comma_v2(transaction_fee_wechat)
            data_list.append(new_transaction_fee_wechat)

            # 取扱高 alipay
            transaction_fee_alipay = i['transaction_fee']['alipay'] if mdr_alipay not in [-1,'-1'] else '-'
            if transaction_fee_alipay != '-':
                transaction_fee_alipay_count += transaction_fee_alipay
            else:
                transaction_fee_wechat_count += 0
            new_transaction_fee_alipay = convert_str_contain_comma_v2(transaction_fee_alipay)
            data_list.append(new_transaction_fee_alipay)

            # 取扱高 wechat+alipay
            transaction_fee_wechat_alipay = i['transaction_fee']['wechat+alipay']
            transaction_fee_wechat_alipay_count += transaction_fee_wechat_alipay
            new_transaction_fee_wechat_alipay = convert_str_contain_comma_v2(transaction_fee_wechat_alipay)
            data_list.append(new_transaction_fee_wechat_alipay)

            # 银行收益 wechat
            agency_profit_wechat = i['agency_profit']['wechat'] if mdr_wechat not in [-1,'-1'] else '-'
            if agency_profit_wechat != '-':
                agency_profit_wechat_count += agency_profit_wechat
            else:
                agency_profit_wechat_count += 0

            new_agency_profit_wechat = convert_str_contain_comma_v2(agency_profit_wechat)
            data_list.append(new_agency_profit_wechat)
            # 银行收益 alipay
            agency_profit_alipay = i['agency_profit']['alipay'] if mdr_alipay not in [-1,'-1'] else '-'
            if agency_profit_alipay != '-':
                agency_profit_alipay_count += agency_profit_alipay
            else:
                agency_profit_alipay_count += 0
            new_agency_profit_alipay = convert_str_contain_comma_v2(agency_profit_alipay)
            data_list.append(new_agency_profit_alipay)
            # 银行收益 wechat+alipay
            agency_profit_wechat_alipay = i['agency_profit']['wechat+alipay']
            agency_profit_wechat_alipay_count += agency_profit_wechat_alipay
            new_agency_profit_wechat_alipay = convert_str_contain_comma_v2(agency_profit_wechat_alipay)
            data_list.append(new_agency_profit_wechat_alipay)
            # print(data_list)
            list_.append(data_list)

    # 获取日期并赋值
    begin_date = stub_data['begin_date']
    end_date = stub_data['end_date']
    new_begin_date = '.'.join(begin_date.split('-'))
    new_end_date = '.'.join(end_date.split('-'))
    new_date = new_begin_date + '~' + new_end_date[5:]

    # print(new_begin_date)
    if new_begin_date == '2019.09.01':

        wb = openpyxl.Workbook()
        ws1 = wb.active

        bd = Side(border_style='thin', color='FF000000')
        ws1.border = Border(left=bd,
                            top=bd,
                            right=bd,
                            bottom=bd)

        # 设置title块的样式
        title_style = NamedStyle(name='title_style')
        title_style.font = Font(color=colors.BLACK)

        bd = Side(border_style='thin', color='FF000000')
        title_style.border = Border(left=bd,
                                    top=bd,
                                    right=bd,
                                    bottom=bd
                                    )

        # horizontal 水平     vertical 垂直
        title_style.alignment = Alignment(horizontal='center',
                                          vertical='center'
                                          )

        # 创建内容块一的样式
        content_style1 = NamedStyle(name='content_style1')
        content_style1.font = Font(color=colors.BLACK)

        bd = Side(border_style='thin', color='FF000000')
        content_style1.border = Border(left=bd,
                                       top=bd,
                                       right=bd,
                                       bottom=bd
                                       )

        content_style1.alignment = Alignment(horizontal='left',
                                             vertical='center'
                                             )

        # 创建内容块二的样式
        content_style2 = NamedStyle(name='content_style2')
        content_style2.font = Font(color=colors.BLACK)

        bd = Side(border_style='thin', color='FF000000')
        content_style2.border = Border(left=bd,
                                       top=bd,
                                       right=bd,
                                       bottom=bd
                                       )

        content_style2.alignment = Alignment(horizontal='right',
                                             vertical='center'
                                             )

        # 创建内容块三的样式
        content_style3 = NamedStyle(name='content_style3')
        content_style3.font = Font(color=colors.BLACK)

        bd = Side(border_style='thin', color='FF000000')
        content_style3.border = Border(left=bd,
                                       top=bd,
                                       right=bd,
                                       bottom=bd
                                       )

        content_style3.alignment = Alignment(horizontal='center',
                                             vertical='center'
                                             )

        # 创建命名样式后，将其注册到工作簿中
        wb.add_named_style(title_style)
        wb.add_named_style(content_style1)
        wb.add_named_style(content_style2)
        wb.add_named_style(content_style3)

        # print(wb._named_styles.names)
        # print(wb._named_styles.names.index('title_style'))


        title_index_list = ['A1', 'B1', 'C1', 'D1', 'E1', 'E2', 'F2', 'G1', 'G2', 'H2',
                            'I2', 'J1', 'J2', 'K2', 'L2', 'M1', 'M2', 'N2', 'O2']
        title_value_list = ['日付', '代理店名', '加盟店名', '店舗数', '加盟店MDR', 'WeChatPay', 'AliPay', '取引件数', 'WeChatPay', 'AliPay',
                            'WeChatPay+AliPay', '取扱高', 'WeChatPay', 'AliPay', 'WeChatPay+AliPay', '銀行收益', 'WeChatPay', 'AliPay',
                            'WeChatPay+AliPay']

        # title块赋值及样式设计
        for t, t_v in zip(title_index_list, title_value_list):
            # print(t, t_v)
            ws1[t].style = title_style
            ws1[t].value = t_v
            # ws1.column_dimensions[t].auto_width = True


        # 合并单元格
        ws1.merge_cells('A1:A2')
        ws1.merge_cells('B1:B2')
        ws1.merge_cells('C1:C2')
        ws1.merge_cells('D1:D2')
        ws1.merge_cells('E1:F1')
        ws1.merge_cells('G1:I1')
        ws1.merge_cells('J1:L1')
        ws1.merge_cells('M1:O1')







        # 获取所有子店铺数
        total_count = stub_data['total_count']
        # print(total_count)

        # 將单元格先设置边框线  然后合并
        for num in range(3, total_count+1):
            bd = Side(border_style='thin', color='FF000000')
            ws1['B' + str(num)].border = Border(
                left=bd,
                top=bd,
                right=bd,
                bottom=bd)
            ws1['A' + str(num)].border = Border(
                left=bd,
                top=bd,
                right=bd,
                bottom=bd)

        # 获取所有店铺数 将 日付 单元格合并
        ws1.merge_cells(start_row=3,
                        start_column=1,
                        end_row=total_count+2,
                        end_column=1)

        # 获取子商铺数 将 代理店名 单元格合并
        # num = 1
        # start_num = 3
        # sub_count = stub_data['data']
        # for sub in sub_count:
        #     count = sub['count']
        #     # print(count)
        #     agency_name = sub['agency_name']
        #     # print(agency_name)
        #     num += count
        #     # print(num)
        #     ws1.merge_cells(start_row=start_num,
        #                     start_column=2,
        #                     end_row=start_num+count,
        #                     end_column=2)
        #     # while count:
        #     #     start_num += count
        #     #     ws1.merge_cells(start_row=start_num,
        #     #                     start_column=2,
        #     #                     end_row=start_num + 1,
        #     #                     end_column=2)
        #     #     # print(start_num)
        #     #     break
        #     # # ws1.merge_cells('B' + str(num) + ':' + 'B' + str(num+1))
        #     # # print(start_num-count)
        #     ws1['B' + str(start_num)].value = agency_name
        #     ws1['B' + str(num)].style = content_style1

        sub_count = stub_data['data']
        start_num = 3
        for sub in sub_count:
            count = sub['count']
            # print(count)
            agency_name = sub['agency_name']
            print(agency_name)
            ws1.merge_cells(start_row=start_num,
                            start_column=2,
                            end_row=start_num+(count - 1),
                            end_column=2)
            ws1['B' + str(start_num)].value = agency_name
            ws1['B' + str(start_num)].style = content_style3
            start_num += count


        # 写入每个加盟店下的所有信息
        for i, item in enumerate(list_):
            # print(list_)
            i = i+3
            # print(i)
            # print('=====')
            for j, val in enumerate(item):
                # print(j)
                ws1.cell(row=i, column=j+3, value=val).style = content_style2
        # for i in list_:
        #     # print(i)
        #     for




        # for n in range(3, total_count+1):
        #     ws1['C' + str(n)].style = content_style1
        #     ws1['B' + str(n)].style = content_style3





        ws1['A3'].style = content_style3
        ws1['A3'].value = new_date
        # 给sheet赋值
        ws1.title = new_date


        # # 设置行高
        ws1.row_dimensions[1].height = 18
        # 设置列宽
        column_name_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
        column_width_list = [20, 30, 30, 18, 15, 15, 15, 15, 20, 15, 15, 20, 15, 15, 20]
        for c, l in zip(column_name_list, column_width_list):
            ws1.column_dimensions[c].width = l


        # 最后一行加入新的一行
        ws1.insert_rows(ws1.max_row + 1)
        row = ws1.max_row + 1

        ws1['A' + str(row)].value = '合計'
        ws1['A' + str(row)].style = content_style1
        ws1['B' + str(ws1.max_row)].style = content_style2
        ws1['C' + str(ws1.max_row)].style = content_style2
        columns_list = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
        for columns in range(3, ws1.max_row+1):
            # print(ws1.max_column)
            for c in columns_list:
                ws1[c + str(columns)].style = content_style2

        ws1['D' + str(row)].value = convert_str_contain_comma_v2(sub_mch_total_count)
        ws1['G' + str(row)].value = convert_str_contain_comma_v2(transaction_total_count_wechat)
        ws1['H' + str(row)].value = convert_str_contain_comma_v2(transaction_total_count_alipay)
        ws1['I' + str(row)].value = convert_str_contain_comma_v2(transaction_total_count_wechat_alipay)
        ws1['J' + str(row)].value = convert_str_contain_comma_v2(transaction_fee_wechat_count)
        ws1['K' + str(row)].value = convert_str_contain_comma_v2(transaction_fee_alipay_count)
        ws1['L' + str(row)].value = convert_str_contain_comma_v2(transaction_fee_wechat_alipay_count)
        ws1['M' + str(row)].value = convert_str_contain_comma_v2(agency_profit_wechat_count)
        ws1['N' + str(row)].value = convert_str_contain_comma_v2(agency_profit_alipay_count)
        ws1['O' + str(row)].value = convert_str_contain_comma_v2(agency_profit_wechat_alipay_count)
        # ws1['D' + str(row)].value = sub_mch_total_count
        # ws1['G' + str(row)].value = transaction_total_count_wechat
        # ws1['H' + str(row)].value = transaction_total_count_alipay
        # ws1['I' + str(row)].value = transaction_total_count_wechat_alipay
        # ws1['J' + str(row)].value = transaction_fee_wechat_count
        # ws1['K' + str(row)].value = transaction_fee_alipay_count
        # ws1['L' + str(row)].value = transaction_fee_wechat_alipay_count
        # ws1['M' + str(row)].value = agency_profit_wechat_count
        # ws1['N' + str(row)].value = agency_profit_alipay_count
        # ws1['O' + str(row)].value = agency_profit_wechat_alipay_count

        # 当前日报表的最后一天为报表标题
        file_name = new_end_date.split('.')
        finally_file_name = '隔週報_Rakuten_' + file_name[0] + file_name[1] + file_name[-1] + '.xlsx'
        wb.save(os.path.join(BASE_DIR, finally_file_name))
        # print(BASE_DIR)
        os.system('cp %s %s' % (os.path.join(BASE_DIR, finally_file_name), os.path.join(COPY_DIR, finally_file_name)))

    else:
        times = datetime.datetime.strftime(datetime.datetime.strptime(begin_date, '%Y-%m-%d') - datetime.timedelta(days=1), '%Y-%m-%d')
        # print(times)
        new_date_time = times.replace('-', '.')
        file_name = new_date_time.split('.')
        finally_file_name = '隔週報_Rakuten_' + file_name[0] + file_name[1] + file_name[-1] + '.xlsx'
        # print(new_date_time)
        wb = load_workbook(os.path.join(BASE_DIR, finally_file_name), read_only=False)
        ws_new = wb.create_sheet(title=new_date, index=0)
        # wa = wb.active
        # print(wa)
        bd = Side(border_style='thin', color='FF000000')
        ws_new.border = Border(left=bd,
                               top=bd,
                               right=bd,
                               bottom=bd)

        # 设置title块的样式
        # title_style = NamedStyle(name='title_style')
        # title_style.font = Font(color=colors.BLACK)
        #
        # bd = Side(border_style='thin', color='FF000000')
        # title_style.border = Border(left=bd,
        #                             top=bd,
        #                             right=bd,
        #                             bottom=bd
        #                             )
        #
        # # horizontal 水平     vertical 垂直
        # title_style.alignment = Alignment(horizontal='center',
        #                                   vertical='center'
        #                                   )
        #
        # # 创建内容块一的样式
        # content_style1 = NamedStyle(name='content_style1')
        # content_style1.font = Font(color=colors.BLACK)
        #
        # bd = Side(border_style='thin', color='FF000000')
        # content_style1.border = Border(left=bd,
        #                                top=bd,
        #                                right=bd,
        #                                bottom=bd
        #                                )
        #
        # content_style1.alignment = Alignment(horizontal='left',
        #                                      vertical='center'
        #                                      )
        #
        # # 创建内容块二的样式
        # content_style2 = NamedStyle(name='content_style2')
        # content_style2.font = Font(color=colors.BLACK)
        #
        # bd = Side(border_style='thin', color='FF000000')
        # content_style2.border = Border(left=bd,
        #                                top=bd,
        #                                right=bd,
        #                                bottom=bd
        #                                )
        #
        # content_style2.alignment = Alignment(horizontal='right',
        #                                      vertical='center'
        #                                      )
        #
        # # 创建内容块三的样式
        # content_style3 = NamedStyle(name='content_style3')
        # content_style3.font = Font(color=colors.BLACK)
        #
        # bd = Side(border_style='thin', color='FF000000')
        # content_style3.border = Border(left=bd,
        #                                top=bd,
        #                                right=bd,
        #                                bottom=bd
        #                                )
        #
        # content_style3.alignment = Alignment(horizontal='center',
        #                                      vertical='center'
        #                                      )

        # 创建命名样式后，将其注册到工作簿中
        # wb.add_named_style(title_style)
        # wb.add_named_style(content_style1)
        # wb.add_named_style(content_style2)
        # wb.add_named_style(content_style3)

        title_index_list = ['A1', 'B1', 'C1', 'D1', 'E1', 'E2', 'F2', 'G1', 'G2', 'H2',
                            'I2', 'J1', 'J2', 'K2', 'L2', 'M1', 'M2', 'N2', 'O2']
        title_value_list = ['日付', '代理店名', '加盟店名', '店舗数', '加盟店MDR', 'WeChatPay', 'AliPay', '取引件数', 'WeChatPay', 'AliPay',
                            'WeChatPay+AliPay', '取扱高', 'WeChatPay', 'AliPay', 'WeChatPay+AliPay', '銀行收益', 'WeChatPay', 'AliPay',
                            'WeChatPay+AliPay']

        title_style_index = wb._named_styles.names.index('title_style')
        content_style1_index = wb._named_styles.names.index('content_style1')
        content_style2_index = wb._named_styles.names.index('content_style2')
        content_style3_index = wb._named_styles.names.index('content_style3')
        # title块赋值及样式设计
        title_style = wb._named_styles.names[title_style_index]
        content_style1 = wb._named_styles.names[content_style1_index]
        content_style2 = wb._named_styles.names[content_style2_index]
        content_style3 = wb._named_styles.names[content_style3_index]
        for t, t_v in zip(title_index_list, title_value_list):
            # print(t, t_v)
            ws_new[t].style = title_style
            ws_new[t].value = t_v
            # ws1.column_dimensions[t].auto_width = True

        # 合并单元格
        ws_new.merge_cells('A1:A2')
        ws_new.merge_cells('B1:B2')
        ws_new.merge_cells('C1:C2')
        ws_new.merge_cells('D1:D2')
        ws_new.merge_cells('E1:F1')
        ws_new.merge_cells('G1:I1')
        ws_new.merge_cells('J1:L1')
        ws_new.merge_cells('M1:O1')

        # 获取所有子店铺数
        total_count = stub_data['total_count']
        # print(total_count)

        # 將单元格先设置边框线  然后合并
        for num in range(3, total_count + 1):
            bd = Side(border_style='thin', color='FF000000')
            ws_new['B' + str(num)].border = Border(
                left=bd,
                top=bd,
                right=bd,
                bottom=bd)
            ws_new['A' + str(num)].border = Border(
                left=bd,
                top=bd,
                right=bd,
                bottom=bd)

        # 获取所有店铺数 将 日付 单元格合并
        ws_new.merge_cells(start_row=3,
                           start_column=1,
                           end_row=total_count + 2,
                           end_column=1)

        # 获取子商铺数 将 代理店名 单元格合并
        new_num = 1
        new_start_num = 3
        sub_count = stub_data['data']
        for sub in sub_count:
            count = sub['count']
            # print(count)
            agency_name = sub['agency_name']
            # print(agency_name)
            new_num += count
            # print(num)
            ws_new.merge_cells(start_row=new_start_num,
                               start_column=2,
                               end_row=new_num + 1,
                               end_column=2)
            while count:
                ws_new.merge_cells(start_row=new_start_num,
                                   start_column=2,
                                   end_row=new_start_num+count-1,
                                   end_column=2)
                # print(start_num)
                break
            # ws1.merge_cells('B' + str(num) + ':' + 'B' + str(num+1))
            # print(start_num-count)
            ws_new['B' + str(new_start_num)].value = agency_name
            ws_new['B' + str(new_start_num)].style = content_style1
            new_start_num = count + new_start_num

        # 写入每个加盟店下的所有信息
        for i, item in enumerate(list_):
            i = i + 3

            for j, val in enumerate(item):
                ws_new.cell(row=i, column=j + 3, value=val).style = content_style2

        for n in range(3, total_count + 1):
            ws_new['C' + str(n)].style = content_style1
            ws_new['B' + str(n)].style = content_style3

        ws_new['A3'].style = content_style3
        ws_new['A3'].value = new_date
        # 给sheet赋值
        ws_new.title = new_date

        # # 设置行高
        ws_new.row_dimensions[1].height = 18
        # 设置列宽
        column_name_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
        column_width_list = [20, 30, 30, 18, 15, 15, 15, 15, 20, 15, 15, 20, 15, 15, 20]
        for c, l in zip(column_name_list, column_width_list):
            ws_new.column_dimensions[c].width = l

        # 最后一行加入新的一行
        ws_new.insert_rows(ws_new.max_row + 1)
        row = ws_new.max_row + 1

        ws_new['A' + str(row)].value = '合計'
        ws_new['A' + str(row)].style = content_style1
        ws_new['B' + str(ws_new.max_row)].style = content_style2
        ws_new['C' + str(ws_new.max_row)].style = content_style2
        columns_list = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
        for columns in range(3, ws_new.max_row + 1):
            for c in columns_list:
                ws_new[c + str(columns)].style = content_style2

        ws_new['D' + str(row)].value = convert_str_contain_comma_v2(sub_mch_total_count)
        ws_new['G' + str(row)].value = convert_str_contain_comma_v2(transaction_total_count_wechat)
        ws_new['H' + str(row)].value = convert_str_contain_comma_v2(transaction_total_count_alipay)
        ws_new['I' + str(row)].value = convert_str_contain_comma_v2(transaction_total_count_wechat_alipay)
        ws_new['J' + str(row)].value = convert_str_contain_comma_v2(transaction_fee_wechat_count)
        ws_new['K' + str(row)].value = convert_str_contain_comma_v2(transaction_fee_alipay_count)
        ws_new['L' + str(row)].value = convert_str_contain_comma_v2(transaction_fee_wechat_alipay_count)
        ws_new['M' + str(row)].value = convert_str_contain_comma_v2(agency_profit_wechat_count)
        ws_new['N' + str(row)].value = convert_str_contain_comma_v2(agency_profit_alipay_count)
        ws_new['O' + str(row)].value = convert_str_contain_comma_v2(agency_profit_wechat_alipay_count)

        # 当前日报表的最后一天为报表标题
        file_name = new_end_date.split('.')
        finally_file_name = '隔週報_Rakuten_' + file_name[0] + file_name[1] + file_name[-1] + '.xlsx'
        wb.save(os.path.join(BASE_DIR, finally_file_name))
        os.system('cp %s %s' % (os.path.join(BASE_DIR, finally_file_name), os.path.join(COPY_DIR, finally_file_name)))




